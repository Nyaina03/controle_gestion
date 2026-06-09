from django.shortcuts import render
import logging
from django.http import JsonResponse
import traceback

from django.http import JsonResponse
import json
import logging
from datetime import datetime
from django.utils import timezone
from .models import DepensesSIIG, HistoriqueDepensesSIIG
from django.shortcuts import get_object_or_404
import logging
from datetime import datetime
from django.shortcuts import get_object_or_404
from .models import DepensesSIIG, HistoriqueDepensesSIIG, Tiers  # Assurez-vous d'importer Tiers
from pta.models import PTA
from  elaborationBudget.models import ModificationSuppressionBudget



import psycopg2


from rest_framework.decorators import api_view



# Create your views here.
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import JsonResponse
from rest_framework import status
from .models import DepensesSIIG
from tiers.models import Tiers
from .serializers import DepensesSIIGSerializer

import pandas as pd
from datetime import datetime
from rest_framework.response import Response
from rest_framework import status
from rest_framework.views import APIView
from .models import DepensesSIIG
from .serializers import DepensesSIIGSerializer
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from .models import DepensesSIIG
from typeOperation.models import TypeOperation
from comptes.models import Comptes
from datetime import datetime
from .models import DepensesSIIG, HistoriqueDepensesSIIG
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt

import psycopg2
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q
from .models import DepensesSIIG, TypeOperation, Comptes
from localites.models import Ville





class DepensesSIIGView(APIView):
    """
    Vue pour afficher toutes les dépenses ou créer une nouvelle dépense.
    """

    def get(self, request):
        depenses = DepensesSIIG.objects.all()
        serializer = DepensesSIIGSerializer(depenses, many=True)
        return Response(serializer.data)

    def post(self, request):
        print("Données reçues:", request.data)

        # Extraction et validation des champs nécessaires
        id_type_operation = request.data.get('id_type_operation')
        annee = request.data.get('annee')
        num_engagement = request.data.get('num_engagement')
        code_programme = request.data.get('code_programme')
        code_tiers = request.data.get('code_tiers')
        id_compte = request.data.get('id_compte')
        objet = request.data.get('objet')
        date_eng = request.data.get('date_eng')
        grande_rubrique = request.data.get('grande_rubrique')
        montant = request.data.get('montant')
        etat = request.data.get('etat')

        # Vérification des champs requis
        if not id_type_operation or not code_programme or not id_compte:
            return Response({
                'error': 'Les champs id_type_operation, code_programme et id_compte sont obligatoires'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Conversion des IDs en entiers
        try:
            id_type_operation = int(id_type_operation)
            code_programme = int(code_programme)
            id_compte = int(id_compte)
        except (ValueError, TypeError):
            return Response({
                'error': 'id_type_operation, code_programme et id_compte doivent être des entiers'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Conversion et validation de la date
        try:
            date_eng = datetime.strptime(date_eng, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return Response({'error': "date_eng doit être au format 'YYYY-MM-DD'"}, status=status.HTTP_400_BAD_REQUEST)

        # Conversion du montant
        try:
            montant = float(montant)
        except (ValueError, TypeError):
            return Response({'error': 'montant doit être un nombre valide'}, status=status.HTTP_400_BAD_REQUEST)

        # Mise à jour des données validées dans la requête
        request.data['id_type_operation'] = id_type_operation
        request.data['annee'] = annee
        request.data['num_engagement'] = num_engagement
        request.data['code_programme'] = code_programme
        request.data['code_tiers'] = code_tiers
        request.data['id_compte'] = id_compte
        request.data['objet'] = objet
        request.data['date_eng'] = date_eng
        request.data['grande_rubrique'] = grande_rubrique
        request.data['montant'] = montant
        request.data['etat'] = etat

        # Sauvegarde dans la base de données
        serializer = DepensesSIIGSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



class DepensesSIIG5ans(APIView):
    """
    Vue pour afficher les dépenses pour une année donnée ou créer une nouvelle dépense.
    """

    def get(self, request):
        # Récupérer l'année passée en paramètre de requête, avec une valeur par défaut
        annee_param = request.query_params.get('annee', None)

        # Si un paramètre 'annee' est passé, filtrer les dépenses par cette année
        if annee_param:
            depenses = DepensesSIIG.objects.filter(annee=annee_param)
        else:
            # Sinon, retourner toutes les dépenses
            depenses = DepensesSIIG.objects.all()

        # Sérialiser les données
        serializer = DepensesSIIGSerializer(depenses, many=True)

        # Retourner la réponse avec les données sérialisées
        return Response(serializer.data)


# Créez un logger
logger = logging.getLogger(__name__)

class ImportDonneeSiigView(APIView):
    """
    Vue pour importer un fichier Excel, matcher les ID avec les tables de référence,
    vérifier les doublons et insérer les données dans la table 'depenses_siig'.
    """
    def post(self, request):
        fichier = request.FILES.get('file')
        if not fichier:
            print("Aucun fichier envoyé.")  # Affichage dans la console du backend
            return Response(
                {'error': 'Aucun fichier envoyé. Veuillez vérifier que le fichier a bien été sélectionné.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Connexion à la base de données PostgreSQL
        conn = psycopg2.connect(
            dbname="controle_gestion",
            user="postgres",
            password="tatamo",
            host="localhost",
            port="5432"
        )
        cursor = conn.cursor()
        print("Connexion à la base de données PostgreSQL réussie.")  # Affichage dans la console du backend

        # Charger le fichier Excel
        try:
            wb = load_workbook(fichier)
            sheet = wb.active
            print("Fichier Excel chargé avec succès.")  # Affichage dans la console du backend
        except Exception as e:
            print(f"Erreur lors du chargement du fichier Excel: {str(e)}")  # Affichage dans la console du backend
            return Response({'error': 'Erreur lors du chargement du fichier Excel.', 'details': str(e)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # En-tête du fichier (noms des colonnes)
        header_row = [cell.value for cell in sheet[1]]
        print(f"En-tête du fichier: {header_row}")  # Affichage dans la console du backend

        # Noms de colonnes attendus dans la base de données
        colonne_attendue = {
            'type_operation': 'id_type_operation',
            'num_engagement': 'num_engagement',
            'code_compte': 'id_compte',
            'date_eng': 'date_eng',
            'montant': 'montant',
            'code_programme': 'code_programme',
            'code_tiers': 'code_tiers',
            'grande_rubrique': 'grande_rubrique',
            'etat': 'etat',
        }

        # Fonction pour récupérer les IDs de la base de données
        def get_id_type_operation(type_operation):
            print(f"Recherche de l'ID pour type_operation: '{type_operation}'")  # Affichage dans la console du backend
            cursor.execute("SELECT id_type_operation FROM type_operation WHERE type_operation = %s", (type_operation,))
            result = cursor.fetchone()
            if result:
                print(f"ID trouvé pour type_operation '{type_operation}': {result[0]}")  # Affichage dans la console du backend
                return result[0]
            else:
                print(f"Aucun ID trouvé pour type_operation: '{type_operation}'")  # Affichage dans la console du backend
            return None


        def get_id_compte(code):
            print(f"Recherche de l'ID pour code_compte: '{code}'")  # Affichage dans la console du backend
            cursor.execute("SELECT id_compte FROM comptes WHERE code = %s", (code,))
            result = cursor.fetchone()
            if result:
                print(f"ID trouvé pour code_compte '{code}': {result[0]}")  # Affichage dans la console du backend
                return result[0]
            else:
                print(f"Aucun ID trouvé pour code_compte: '{code}'")  # Affichage dans la console du backend
            return None


        def get_id_tiers(code_tiers):
            print(f"Recherche de l'ID pour code_tiers: '{code_tiers}'")  # Affichage dans la console du backend
            cursor.execute("SELECT id_tiers FROM tiers WHERE id_tiers = %s", (code_tiers,))
            result = cursor.fetchone()
            if result:
                print(f"ID trouvé pour code_tiers '{code_tiers}': {result[0]}")  # Affichage dans la console du backend
                return result[0]
            else:
                print(f"Aucun ID trouvé pour code_tiers: '{code_tiers}'")  # Affichage dans la console du backend
            return None


        # Initialiser les variables pour le suivi
        erreurs_mapping = []
        lignes_valides = []

        try:
            # Lire les lignes du fichier Excel
            for row in sheet.iter_rows(min_row=2, values_only=True):
                ligne_donnee = {}

                # Récupérer les valeurs des colonnes en fonction des noms
                for col_name, db_column in colonne_attendue.items():
                    if col_name in header_row:
                        column_index = header_row.index(col_name)
                        value = row[column_index]

                        # Associer la valeur au champ de la base de données
                        if db_column == 'id_type_operation':
                            value = get_id_type_operation(value)
                            if not value:
                                print(f"Ligne invalide (manque id_type_operation): {row}")  # Affichage dans la console du backend
                                erreurs_mapping.append({'ligne': row, 'erreur': f'Manque id_type_operation pour {value}'})
                                continue

                        elif key == 'code_programme':
                            if isinstance(value, list) and len(value) == 1:
                                setattr(depense, key, value[0])  # Extraire la première valeur
                            elif isinstance(value, str):
                                setattr(depense, key, value)
                            else:
                                return JsonResponse(
                                    {'success': False, 'message': f"Valeur invalide pour {key}: {value}"},
                                    status=status.HTTP_400_BAD_REQUEST
                                )

                        elif db_column == 'id_compte':
                            value = get_id_compte(value)
                            if not value:
                                print(f"Ligne invalide (manque id_compte): {row}")  # Affichage dans la console du backend
                                erreurs_mapping.append({'ligne': row, 'erreur': f'Manque id_compte pour {value}'})
                                continue
                        elif db_column == 'date_eng':
                            try:
                                if isinstance(value, datetime):
                                    # Si c'est déjà un objet datetime, utilisez-le directement
                                    value = value.date()
                                else:
                                    # Sinon, effectuez la conversion avec strptime
                                    value = datetime.strptime(value, "%d/%m/%Y").date()
                            except Exception as e:
                                print(f"Erreur de conversion pour {col_name} (ligne: {row}): {str(e)}")  # Affichage dans la console du backend
                                erreurs_mapping.append({'ligne': row, 'erreur': f'Conversion invalide pour {col_name}: {str(e)}'})
                                continue

                        elif db_column == 'montant':
                            try:
                                value = float(str(value).replace(',', '.'))
                            except Exception as e:
                                print(f"Erreur de conversion pour {col_name} (ligne: {row}): {str(e)}")  # Affichage dans la console du backend
                                erreurs_mapping.append({'ligne': row, 'erreur': f'Conversion invalide pour {col_name}: {str(e)}'})
                                continue
                        ligne_donnee[db_column] = value

                # Afficher les données après le matching
                print(f"Données après le matching: {ligne_donnee}")  # Affichage dans la console du backend

                # Vérification de la validité des données
                if ligne_donnee.get('id_type_operation') and ligne_donnee.get('id_compte'):
                    print(f"Ligne valide à insérer: {ligne_donnee}")  # Affichage dans la console du backend
                    lignes_valides.append(ligne_donnee)
                else:
                    print(f"Ligne invalide (manque id_type_operation ou id_compte): {ligne_donnee}")  # Affichage dans la console du backend
                    erreurs_mapping.append({'ligne': row, 'erreur': 'Correspondance manquante pour certains champs'})

            # Insertion des données dans la base
            for ligne in lignes_valides:
                serializer = DepensesSIIGSerializer(data=ligne)
                if serializer.is_valid():
                    serializer.save()
                    print(f"Ligne insérée avec succès: {ligne}")  # Affichage dans la console du backend
                else:
                    erreurs_mapping.append({'ligne': ligne, 'erreur': serializer.errors})
                    print(f"Erreur d'insertion pour {ligne}: {serializer.errors}")  # Affichage dans la console du backend

            # Fermer la connexion à la base de données PostgreSQL
            cursor.close()
            conn.close()
            print("Connexion à la base de données fermée.")  # Affichage dans la console du backend

            return Response({
                'message': f'{len(lignes_valides)} nouvelles lignes insérées avec succès.',
                'erreurs': erreurs_mapping
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            # Si une erreur se produit, on ferme la connexion et on retourne l'erreur
            conn.rollback()
            cursor.close()
            conn.close()
            print(f"Erreur pendant le traitement du fichier: {str(e)}")  # Affichage dans la console du backend
            return Response({'error': 'Erreur pendant le traitement du fichier.', 'details': str(e)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DepensesSIIGDetailView(APIView):

    def get(self, request, id_type_operation, annee, num_engagement, *args, **kwargs):
        try:
            # Récupérer les dépenses en filtrant selon les paramètres d'URL
            queryset = DepensesSIIG.objects.all()

            # Appliquer les filtres sur les paramètres d'URL
            if id_type_operation:
                queryset = queryset.filter(id_type_operation=id_type_operation)
            if annee:
                queryset = queryset.filter(annee=annee)
            if num_engagement:
                queryset = queryset.filter(num_engagement=num_engagement)

            # Utiliser select_related pour effectuer une jointure avec la table comptes et récupérer le code
            queryset = queryset.select_related('id_compte')

            # Debug : Afficher la requête SQL (à enlever en production)
            print(queryset.query)

            # Vérifier si des résultats existent
            if not queryset.exists():
                return Response({"error": "Aucune dépense trouvée avec les critères donnés."},
                                status=status.HTTP_404_NOT_FOUND)

            # Sérialisation des données
            # Nous ajoutons l'id_depense_siig et le code du compte à chaque élément retourné
            result_data = []
            for depense in queryset:
                # Ajouter l'id et le code du compte à chaque objet de dépense
                depense_data = DepensesSIIGSerializer(depense).data
                depense_data['id_depense_siig'] = depense.id_depense_siig  # Assurez-vous que ce champ existe dans le modèle
                depense_data['code_compte'] = depense.id_compte.code  # Assurez-vous que `id_compte` a un champ `code`
                result_data.append(depense_data)

            return Response(result_data, status=status.HTTP_200_OK)

        except ValueError as ve:
            return Response({"error": f"Erreur de conversion de type: {str(ve)}"},
                            status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": f"Erreur lors de la récupération des dépenses: {str(e)}"},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def put(self, request, id_depense_siig, format=None):
        try:
            depense = DepensesSIIG.objects.get(id_depense_siig=id_depense_siig)
            serializer = DepensesSIIGSerializer(depense, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except DepensesSIIG.DoesNotExist:
            return Response({"error": "Dépense non trouvée"}, status=status.HTTP_404_NOT_FOUND)

    
    def delete(self, request, id_depense_siig, format=None):
        try:
            depense = DepensesSIIG.objects.get(id_depense_siig=id_depense_siig)
            depense.delete()
            return Response({"message": "Dépense supprimée avec succès"}, status=status.HTTP_204_NO_CONTENT)
        except DepensesSIIG.DoesNotExist:
            return Response({"error": "Dépense non trouvée"}, status=status.HTTP_404_NOT_FOUND)


class UpdateDepenseView(APIView):
    #def put(self, request, id):  # Utilisez 'id' si c'est défini dans la route
    #    try:
    #        depense_siig = DepensesSIIG.objects.get(pk=id)
    #    except DepensesSIIG.DoesNotExist:
    #        return Response({'error': 'Dépense non trouvée.'}, status=status.HTTP_404_NOT_FOUND)

        # serializer = DepensesSIIGSerializer(depense_siig, data=request.data)
        #if serializer.is_valid():
        #    serializer.save()
        #    return Response({'success': True, 'data': serializer.data}, status=status.HTTP_200_OK)
        #return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
    def put(self, request, id):
        try:
            # Récupérer la dépense existante (depense) par son ID
            depense = get_object_or_404(DepensesSIIG, pk=id)
            print("Dépense trouvée:", depense)

            # Récupérer les données du corps de la requête PUT
            new_data = json.loads(request.body)
            print("Données reçues pour mise à jour:", new_data)

            # Validation initiale des champs obligatoires
            required_fields = ['id_compte', 'id_type_operation']
            for field in required_fields:
                if field in new_data and not new_data[field]:
                    return JsonResponse(
                        {'success': False, 'message': f"Le champ {field} est obligatoire et ne peut être vide."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Récupérer l'instance de Tiers en utilisant code_tiers pour la clé étrangère
            tiers_instance = depense.code_tiers  # Accès à l'objet Tiers via la clé étrangère 'code_tiers'

            # Sauvegarder l'historique avant la mise à jour
            HistoriqueDepensesSIIG.objects.create(
                id_depense_siig=depense.id_depense_siig,
                id_type_operation=depense.id_type_operation.id_type_operation if depense.id_type_operation else None,
                annee=depense.annee,
                num_engagement=depense.num_engagement,
                code_programme=depense.code_programme,
                code_tiers=tiers_instance,  # On assigne l'instance Tiers ici
                id_compte=depense.id_compte.id_compte if depense.id_compte else None,
                objet=depense.objet,
                date_eng=depense.date_eng,
                grande_rubrique=depense.grande_rubrique,
                montant=depense.montant,
                etat=depense.etat,
                date_modification=datetime.now(),
            )
            print("Historique sauvegardé.")

            # Mise à jour de la dépense
            serializer = DepensesSIIGSerializer(depense, data=new_data)  # Correction ici, il faut utiliser new_data
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            print("Erreur interne:", str(e))
            return JsonResponse({'success': False, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


    def historique(self, request, id):  # 'id' est le paramètre URL
        try:
            # Récupérer la dépense existante (depense) par son ID
            depense = get_object_or_404(DepensesSIIG, pk=id)
            print("Dépense trouvée:", depense)

            # Récupérer les données du corps de la requête PUT
            new_data = json.loads(request.body)
            print("Données reçues pour mise à jour:", new_data)

            # Validation initiale des champs obligatoires
            required_fields = ['id_compte', 'id_type_operation']
            for field in required_fields:
                if field in new_data and not new_data[field]:
                    return JsonResponse(
                        {'success': False, 'message': f"Le champ {field} est obligatoire et ne peut être vide."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Sauvegarder l'historique avant la mise à jour
            HistoriqueDepensesSIIG.objects.create(
                id_depense_siig=depense.id_depense_siig,
                id_type_operation=depense.id_type_operation.id_type_operation if depense.id_type_operation else None,
                annee=depense.annee,
                num_engagement=depense.num_engagement,
                code_programme=depense.code_programme,
                code_tiers=depense.code_tiers.id_tiers if depense.code_tiers else None,
                id_compte=depense.id_compte.id_compte if depense.id_compte else None,
                objet=depense.objet,
                date_eng=depense.date_eng,
                grande_rubrique=depense.grande_rubrique,
                montant=depense.montant,
                etat=depense.etat,
                date_modification=now(),
            )
            print("Historique sauvegardé.")

        except Exception as e:
            print("Erreur interne:", str(e))
            return JsonResponse({'success': False, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


#changement d'etat
class ChangementEtatView(APIView):
    def put(self, request, pk):
        try:
            depense = DepensesSIIG.objects.get(pk=pk)
        except DepensesSIIG.DoesNotExist:
            return Response({'error': 'engagement not found'}, status=status.HTTP_404_NOT_FOUND)

        serializer = DepensesSIIGSerializer(depense, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class VisualisationEngagements(APIView):
    def get(self, request, *args, **kwargs):
        # Récupération de l'année depuis les paramètres d'URL
        annee = kwargs.get('annee', None)
        
        if not annee:
            return Response({"detail": "L'année est requise"}, status=status.HTTP_400_BAD_REQUEST)

        # Récupérer les engagements pour l'année spécifiée
        engagements = DepensesSIIG.objects.filter(annee=annee).select_related('code_tiers', 'id_compte')

        # Initialisation d'une liste vide pour stocker les résultats
        engagement_data = []

        for engagement in engagements:
            # Récupérer les informations du tiers
            tiers = engagement.code_tiers if engagement.code_tiers else None
            compte = engagement.id_compte if engagement.id_compte else None

            # Assurez-vous que 'tiers' et 'compte' existent et récupérez les données nécessaires
            tiers_nom = tiers.nom if tiers else "Non renseigné"
            tiers_id = tiers.id_tiers if tiers else None  # Extraire l'ID du tiers
            compte_code = compte.code if compte else "Non renseigné"

            # Récupérer l'ID de la ville du tiers si tiers existe
            ville_nom = "Non renseignée"
            if tiers and tiers.id_ville:
                ville = tiers.id_ville  # Utilisez l'ID directement
                ville_nom = ville.nom_ville if ville else "Ville non trouvée"  # Utiliser le nom de la ville pour l'affichage

            # Ajouter l'engagement aux données
            engagement_data.append({
                "numEng": engagement.num_engagement,
                "prog": engagement.code_programme,
                "codeTiers": tiers_id,  # Utiliser l'ID du tiers, pas l'objet entier
                "nomTiers": tiers_nom,
                "compte": compte_code,  # Utiliser 'compte.code' pour accéder au champ
                "objet": engagement.objet,
                "date": engagement.date_eng,
                "type": engagement.etat,
                "grandeRubrique": engagement.grande_rubrique,
                "montant": engagement.montant,
                "etat": engagement.etat,
                "dateVisa": engagement.date_eng,
                "dateOs": engagement.date_eng,
                "marche": engagement.objet,
                "typeEng": engagement.etat,
                "codePta": engagement.code_programme,
                "ville": ville_nom  # Retourner le nom de la ville
            })

        return Response(engagement_data, status=status.HTTP_200_OK)



class EtatEngagementParCompte(APIView):
    def get(self, request, id_compte, annee):
        if not id_compte or not annee:
            return Response(
                {"error": "Les paramètres 'id_compte' et 'annee' sont requis."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Récupération des engagements dans DepensesSIIG pour l'année et le compte donnés
            engagements = (
                DepensesSIIG.objects.filter(annee=annee, id_compte=id_compte)
                .select_related("id_compte")  # Permet de récupérer les données du modèle comptes
                .values(
                    "date_eng",
                    "num_engagement",
                    "objet",
                    "montant",
                    "etat",
                    "code_tiers",
                    "code_programme",
                    "id_compte__code"  # Utiliser le champ correct "code" depuis comptes
                )
            )

            # Récupération des numéros PTA liés aux engagements
            ref_dos_list = [engagement["num_engagement"] for engagement in engagements]
            pta_data = {
                pta.ref_dos: {
                    "code_programme": pta.code_programme,
                    "libelle": pta.libelle
                }
                for pta in PTA.objects.filter(ref_dos__in=ref_dos_list)
            }

            # Récupération des noms des fournisseurs
            code_tiers_list = [engagement["code_tiers"] for engagement in engagements if engagement["code_tiers"]]
            fournisseurs = {
                tiers.id_tiers: tiers.nom or tiers.raison_social
                for tiers in Tiers.objects.filter(id_tiers__in=code_tiers_list)#, id_type_tiers=2#)
            }

            # Préparer la réponse finale
            resultats = []
            for engagement in engagements:
                num_engagement = engagement["num_engagement"]
                code_tiers = engagement["code_tiers"]
                code_compte = engagement.get("id_compte__code", "Non renseigné")  # Récupérer le code du compte

                resultats.append({
                    "dateEng": engagement["date_eng"] or "Non renseigné",
                    "compte": code_compte,  # Inclure le code du compte dans la réponse
                    "numEng": num_engagement,
                    "fournisseurs": fournisseurs.get(code_tiers, "Non renseigné"),
                    "objet": engagement["objet"] or pta_data.get(num_engagement, {}).get("libelle", "Non renseigné"),
                    "montant": float(engagement["montant"] or 0.0),
                    "etat": engagement["etat"] or "Non renseigné",
                    "numPta": pta_data.get(num_engagement, {}).get("code_programme", "Non renseigné"),
                })

            return Response(resultats, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Une erreur s'est produite : {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class EtatEngagementParTiers(APIView):
    def get(self, request, id_compte, annee):
        if not id_compte or not annee:
            return Response(
                {"error": "Les paramètres 'id_compte' et 'annee' sont requis."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Récupération des engagements pour l'année et le compte donnés
            engagements = (
                DepensesSIIG.objects.filter(annee=annee, id_compte=id_compte)
                .select_related("id_compte")
                .values(
                    "date_eng",
                    "num_engagement",
                    "objet",
                    "montant",
                    "etat",
                    "code_tiers",
                    "code_programme",
                    "id_compte__code"
                )
            )

            # Calculer le nombre de paiements effectués (en fonction des attachements)
            paiements_effectues = {}
            for engagement in engagements:
                num_engagement = engagement["num_engagement"]

                # Récupérer les modifications pour l'engagement (filtrées par id_compte et num_marche)
                modifications = ModificationSuppressionBudget.objects.filter(
                    id_compte=id_compte, num_marche=num_engagement
                )

                # Calculer le nombre d'attachements non nuls pour chaque modification
                nb_paiements = 0
                for modification in modifications:
                    # Vérifier chaque attach et compter les non nuls
                    for i in range(1, 21):  # de attach1 à attach20
                        attach_value = getattr(modification, f"attach{i}", None)
                        if attach_value not in [None, 0]:
                             nb_paiements += 1

                paiements_effectues[num_engagement] = nb_paiements


            # Récupération des numéros PTA et fournisseurs
            ref_dos_list = [engagement["num_engagement"] for engagement in engagements]
            pta_data = {
                pta.ref_dos: {
                    "code_programme": pta.code_programme,
                    "libelle": pta.libelle
                }
                for pta in PTA.objects.filter(ref_dos__in=ref_dos_list)
            }

            code_tiers_list = [engagement["code_tiers"] for engagement in engagements if engagement["code_tiers"]]
            fournisseurs = {
                tiers.id_tiers: tiers.nom or tiers.raison_social
                for tiers in Tiers.objects.filter(id_tiers__in=code_tiers_list)
            }

            # Préparer la réponse finale
            resultats = []
            for engagement in engagements:
                num_engagement = engagement["num_engagement"]
                code_tiers = engagement["code_tiers"]
                code_compte = engagement.get("id_compte__code", "Non renseigné")

                resultats.append({
                    "dateEng": engagement["date_eng"] or "Non renseigné",
                    "compte": code_compte,
                    "numEng": num_engagement,
                    "code_tiers": code_tiers,
                    "fournisseurs": fournisseurs.get(code_tiers, "Non renseigné"),
                    "objet": engagement["objet"] or pta_data.get(num_engagement, {}).get("libelle", "Non renseigné"),
                    "montant": float(engagement["montant"] or 0.0),
                    "etat": engagement["etat"] or "Non renseigné",
                    "numPta": pta_data.get(num_engagement, {}).get("code_programme", "Non renseigné"),
                    "nbPaiements": paiements_effectues.get(num_engagement, 0)
                })

            return Response(resultats, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Une erreur s'est produite : {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SuiviDesEngagements(APIView):
    def get(self, request, id_compte, etat_engagement, annee):
        if not id_compte or not annee:
            return Response(
                {"error": "Les paramètres 'id_compte' 'etat_engagement'et 'annee' sont requis."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Récupération des engagements pour l'année et le compte donnés
            engagements = (
                DepensesSIIG.objects.filter(annee=annee, etat=etat_engagement, id_compte=id_compte)
                .select_related("id_compte")
                .values(
                    "date_eng",
                    "num_engagement",
                    "objet",
                    "montant",
                    "etat",
                    "code_tiers",
                    "code_programme",
                    "id_compte__code"
                )
            )
            resultats = []

            resultats.append(engagements)

            return Response(resultats, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Une erreur s'est produite : {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )